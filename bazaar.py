import requests, time, json, os, platform, subprocess, threading
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Font, NamedStyle, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo
from collections import defaultdict

# URL de l'API du bazaar
API_URL = "https://api.hypixel.net/skyblock/bazaar"

# Variable partagée
user_pressed = False


# Fonction mise à jour pour enregistrer ou modifier le fichier Excel
def save_to_excel(data, filename="bazaar_data.xlsx"):
    if os.path.exists(filename):
        os.remove(filename)  # Supprime l'ancien fichier pour tout réécrire

    wb = Workbook()
    sheet = wb.active
    sheet.title = "Bazaar Analysis"

    # Ajouter les en-têtes
    headers = ["Item", "Buy Price", "Sell Price", "Profit", "Supply & Demand", "Profit per coin", "Craftable"]
    sheet.append(headers)

    # Ajouter un style pour les nombres avec un séparateur des milliers
    number_style = NamedStyle(name="number_style", number_format="#,##0.00")

    # Ajouter les données dans le fichier Excel
    for row, values in enumerate(data, start=2):  # Commence à la ligne 2 (après les en-têtes)
        item_id = values["item"]
        item_name = values["name"]  # Nom de l'objet
        wiki_url = values.get("wiki", "")  # URL du wiki, si disponible

        buy_price = values["buy_price"]
        sell_price = values["sell_price"]
        craftable = values["craft"]["craftable"]
        craft_profit = values["craft"]["craft_profit"]
        materials = values["craft"]["materials"]
        profit = values["profit"]
        buy_volume = values["buy_volume"]
        sell_volume = values["sell_volume"]

        # Ajouter les valeurs dans les colonnes
        cell_name = sheet.cell(row=row, column=1, value=item_name)
        if wiki_url and wiki_url != None:
            cell_name.hyperlink = wiki_url
            cell_name.font = Font(color="0000FF", underline="single")
        cell_name.alignment = Alignment(horizontal="center")

        sheet.cell(row=row, column=2, value=buy_price).style = number_style
        sheet.cell(row=row, column=3, value=sell_price).style = number_style
        sheet.cell(row=row, column=4, value=profit).style = number_style
        sheet.cell(row=row, column=5, value=f"{buy_volume}/{sell_volume}")
        sheet.cell(row=row, column=6, value=profit/buy_price).style = number_style
        sheet.cell(row=row, column=7, value=craftable)

    # Ajouter un tableau Excel (avec filtres)
    data_range = f"A1:G{len(data) + 1}"  # La plage de données pour le tableau
    table = Table(displayName="BazaarTable", ref=data_range)

    # Style du tableau
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    sheet.add_table(table)

    # Ajuster la largeur des colonnes
    for col in sheet.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[col_letter].width = max_length + 4
        # === FEUILLE 2 : SIMULATEUR DE FLIP ===
    sim_sheet = wb.create_sheet("Flip Simulator")

    # Titre
    sim_sheet["A1"] = "Simulateur de Flip"
    sim_sheet["A1"].font = Font(bold=True, size=14)

    # Instructions
    sim_sheet["A3"] = "Sélectionne un item à flipper :"
    sim_sheet["A5"] = "Entrez votre budget en coins :"
    sim_sheet["A7"] = "Quantité possible à acheter :"
    sim_sheet["A8"] = "Profit total estimé :"
    sim_sheet["A9"] = "Prix unitaire d'achat:"
    sim_sheet["A10"] = "Prix unitaire de vente:"

    # Cellule pour sélection d'item
    sim_sheet["B3"] = ""  # la cellule contenant la liste déroulante

    # Liste déroulante (basée sur les noms d'items dans la première feuille)
    item_count = len(data)
    item_range = f"'Bazaar Analysis'!A2:A{item_count+1}"
    dv = DataValidation(type="list", formula1=item_range, showDropDown=False, allowBlank=False)
    sim_sheet.add_data_validation(dv)
    dv.add(sim_sheet["B3"])

    # Cellule pour entrée de budget
    sim_sheet["B5"] = 1000000  # valeur par défaut
    sim_sheet["B5"].number_format = "#,##0"

    # Recherche du prix d’achat correspondant à l’item choisi
    # On suppose que l'item choisi est en B3 et le budget en B5
    # Formule Excel pour retrouver le prix d'achat
    sim_sheet["B7"].value = '=IFERROR(ROUNDDOWN(B5 / VLOOKUP(B3, \'Bazaar Analysis\'!A:F, 2, FALSE), 0), "Erreur")'
    sim_sheet["B8"].value = '=IFERROR(B7 * VLOOKUP(B3, \'Bazaar Analysis\'!A:F, 4, FALSE), "Erreur")'
    sim_sheet["B9"].value = '=IFERROR(VLOOKUP(B3, \'Bazaar Analysis\'!A:F, 2, FALSE), "Erreur")'
    sim_sheet["B10"].value = '=IFERROR(VLOOKUP(B3, \'Bazaar Analysis\'!A:F, 3, FALSE), "Erreur")'


    # Appliquer style aligné + gras sur les titres de résultat
    for cell in ["A7", "A8", "A9", "A10"]:
        sim_sheet[cell].font = Font(bold=True)
        sim_sheet[cell].alignment = Alignment(horizontal="left")
    for cell in ["B7", "B8", "B9", "B10"]:
        sim_sheet[cell].number_format = "#,##0"
    
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    sim_sheet["B3"].fill = yellow_fill
    sim_sheet["B5"].fill = yellow_fill

    # Ajuster les colonnes
    for col in sim_sheet.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        sim_sheet.column_dimensions[col_letter].width = max_len + 4


    # Sauvegarder le fichier
    wb.save(filename)

# Charger les données depuis le fichier crafts.json
def load_craft_data(craft_filename="crafts.json"):
    try:
        with open(craft_filename, encoding="utf-8") as file:
            return json.load(file)

    except FileNotFoundError:
        print("Erreur", f"Le fichier {craft_filename} est introuvable.")
        return {}
    except json.JSONDecodeError:
        print("Erreur", f"Erreur lors de la lecture du fichier {craft_filename}.")
        return {}
    
def get_bazaar_data():
    try:
        response = requests.get(API_URL)
        response.raise_for_status()  # Vérifie si la requête a réussi
        return response.json()
    except requests.exceptions.RequestException as e:
        print("Erreur", f"Erreur de connexion à l'API du bazaar : {e}")
        return None

def parse_ingredients(recipe):
    totals = defaultdict(int)
    for value in recipe.values():
        if isinstance(value, str) and ":" in value:
            item_id, amount = value.split(":")
            totals[item_id] += int(amount)

    # This returns a dictionary with item_id as keys and their corresponding amounts as values
    return {k: v for k, v in totals.items()}


def load_crafts(craft_data, bazaar_data):
    possible_crafts_bazaar = {
            item:      
            {
                "item_id": item,
                "ingredients": parse_ingredients(craft_data[item]["recipe"])
            }
            for item in craft_data
            if item in bazaar_data["products"] and "recipe" in craft_data[item]
    }

    with open(f"Ingredients.json", "w", encoding="utf-8") as file:
        json.dump(possible_crafts_bazaar, file, indent=4)
    return possible_crafts_bazaar

def calculate_profit(bazaar_data, craft_data, ingredients):
    profits = []
    craft = {}
    item_craft_profit = 0
    if not ingredients:
        print("Aucun ingredients")
        
    if not bazaar_data:
        print("Aucun bazaar")
        
    if not craft_data:
        print("Aucun craft")
        
    for item_id, product in bazaar_data["products"].items():
        if item_id in ingredients:
            item_craft_profit = 0
            for ingredient, quantity in ingredients[item_id]["ingredients"].items():
                if ingredient in bazaar_data["products"] and bazaar_data["products"][ingredient]["buy_summary"] and bazaar_data["products"][ingredient]["sell_summary"]:
                    buy_price = bazaar_data["products"][ingredient]["sell_summary"][0]["pricePerUnit"]
                    sell_price = bazaar_data["products"][ingredient]["buy_summary"][0]["pricePerUnit"]
                    if (sell_price / buy_price - 1)*100 > 80 and (sell_price - buy_price) > 100:
                        item_craft_profit = 0
                        break
                    profit = round(sell_price - buy_price, 1)
                    item_craft_profit += profit*quantity
            craft = {
                "craftable": True if item_craft_profit > 0 else False,
                "craft_profit": item_craft_profit,
                "materials": {
                    craft_data.get(ingredient, {}).get("name", ingredient): quantity
                    for ingredient, quantity in ingredients[item_id]["ingredients"].items()
                }
            }
            print(f"Item: {item_id}, Profit total du craft: {round(item_craft_profit):,}\n".replace(",", " ") if item_craft_profit > 0 else f"Profit trop grand pour le craft de {item_id}\n")
        else:
            craft = {
                "craftable": False,
                "craft_profit": 0,
                "materials": {}
            }
        # Vérifier si les données d'achat/vente sont disponibles
        if product["buy_summary"] and product["sell_summary"]:
            buy_price = product["sell_summary"][0]["pricePerUnit"]
            sell_price = product["buy_summary"][0]["pricePerUnit"]

            if buy_price == 0 or sell_price == 0:
                continue  # Si les prix sont à 0, on ignore cet item

            profit = round(sell_price - buy_price, 1)

            # Chercher le nom et l'ID wiki dans craft.json
            item_name = craft_data.get(item_id, {}).get("name", item_id)
            if not item_name == item_id:
                wiki_url = f"https://wiki.hypixel.net/{item_name}"  # Lien vers le wiki
            else:
                wiki_url = None
                
            # Ajouter les données au résultat
            profits.append({
                'item': item_id,
                'name': item_name,
                'buy_price': buy_price,
                'sell_price': sell_price,
                'profit': profit,
                'craft': craft,
                'sell_volume': product["sell_summary"][0]["amount"],
                'buy_volume': product["buy_summary"][0]["amount"],
                'wiki': wiki_url  # Lien vers le wiki
            })
    with open(f"Profits.json", "w", encoding="utf-8") as file:
        json.dump(profits, file, indent=4)
    return profits

# Fonction pour exécuter l'analyse et afficher les résultats
def run_analysis():
    # Charger les données craft
    craft_data = load_craft_data()
    if not craft_data:
        return []

    # Récupérer les données du bazaar
    bazaar_data = get_bazaar_data()
    if not bazaar_data:
        return []

    # Calculer les profits
    results = calculate_profit(bazaar_data, craft_data, load_crafts(craft_data, bazaar_data))

    # Afficher les résultats dans la console (ou GUI)
    for result in results:
        print(f"Item: {result['item']}, Profit: {result['profit']}")
        print(f"Wiki: {result['wiki'] if result['wiki'] != None else 'N/A'}")

    return results

def wait_for_input():
    global user_pressed
    input("Appuyez sur Entrée pour garder la console ouverte...\n")
    user_pressed = True

def countdown(seconds=5):
    global user_pressed

    # Lancer un thread pour détecter l'entrée utilisateur
    input_thread = threading.Thread(target=wait_for_input)
    input_thread.daemon = True
    input_thread.start()
    time.sleep(0.1)

    for i in range(seconds, 0, -1):
        if user_pressed:
            break
        print(f"Fermeture dans {i} seconde(s)...")
        time.sleep(1)

    if user_pressed:
        print("Console maintenue ouverte.")
        input("Appuyez sur Entrée pour quitter...")
    else:
        print("Temps écoulé. Fermeture.")

def open_file(filename):
    if platform.system() == "Windows":
        os.startfile(filename)
    elif platform.system() == "Darwin":  # macOS
        subprocess.call(["open", filename])
    else:  # Linux
        subprocess.call(["xdg-open", filename])

# Configuration de l'interface utilisateur avec Tkinter
def start():
    results = run_analysis()
    if not results:
        print("Aucun résultat. Aucune donnée à sauvegarder.")
        input("\n\nAppuie sur Entrée pour quitter...")
        return

    filename = "bazaar_data.xlsx"

    try:
        save_to_excel(results, filename)
        print("Succès : Données sauvegardées dans", filename)
        open_file(filename)
        countdown()
    except Exception as e:
        print("Erreur : Une erreur s'est produite :", e)
        input("\n\nAppuie sur Entrée pour quitter...")

# Lancer le script
start()